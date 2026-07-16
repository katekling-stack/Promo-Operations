"""The master workbook generator produces valid dropdowns sourced from config."""

from __future__ import annotations

import importlib.util
from pathlib import Path

from openpyxl import load_workbook

from promo_ops.config import REPO_ROOT

_spec = importlib.util.spec_from_file_location(
    "build_template_workbook", REPO_ROOT / "scripts" / "build_template_workbook.py")
bw = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(bw)


def test_dropdown_sources_match_config():
    src = bw.dropdown_sources()
    assert "USA" in src["region"]
    assert "Paramount + - USA" in src["campaign name"]
    assert src["content type"] == ["show", "movie"]
    assert "pluto" in src["video domination"]
    assert "hpto" in src["takeover"]
    assert "paramount_plus_domestic" in src["brand"]


def test_workbook_builds_with_dropdowns(tmp_path: Path):
    out = bw.build(tmp_path / "wb.xlsx")
    wb = load_workbook(out)
    assert wb.sheetnames == ["Plan", "Targeting", "_Lists"]
    assert wb["_Lists"].sheet_state == "hidden"
    # 6 constrained fields (region, campaign, content type, VD, takeover, brand)
    # + one Yes/No dropdown per Products toggle.
    expected = 6 + len(bw._product_toggle_labels())
    assert len(wb["Plan"].data_validations.dataValidation) == expected
