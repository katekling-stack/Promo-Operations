"""Generate the master campaign-plan workbook (.xlsx) from the CSV templates + config.

The workbook is what planners actually copy per campaign: the same Plan + Targeting
layout as templates/campaign-plan/*.csv, but with data-validation **dropdowns** on the
constrained fields (Region, Campaign, Content Type, Video Domination, Takeover, Brand)
so invalid values can't be typed. Dropdown values are sourced from the YAML config, so
this stays in sync with what the builder accepts.

Run:  python scripts/build_template_workbook.py
Out:  templates/campaign-plan/Campaign-Plan-Template.xlsx
Upload that file to Google Drive and "Open with Google Sheets" to get the shared master.
"""

from __future__ import annotations

import csv
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

REPO = Path(__file__).resolve().parents[1]
TEMPLATE_DIR = REPO / "templates" / "campaign-plan"
CONFIG = REPO / "config"


def _yaml(name: str) -> dict:
    with (CONFIG / name).open(encoding="utf-8") as fh:
        return yaml.safe_load(fh) or {}


def _csv_rows(name: str) -> list[list[str]]:
    with (TEMPLATE_DIR / name).open(encoding="utf-8", newline="") as fh:
        return list(csv.reader(fh))


def _product_toggle_labels() -> list[str]:
    """The Products section toggle labels (from the shared PRODUCT_TOGGLES map)."""
    from promo_ops.integrations.gsheets import PRODUCT_TOGGLES
    return list(PRODUCT_TOGGLES)


def dropdown_sources() -> dict[str, list[str]]:
    """field-label (lowercased) -> allowed values, sourced from config."""
    regions = list(_yaml("regions.yaml").get("regions", {}))
    brands = _yaml("brands.yaml").get("brands", {})
    campaigns = [b.get("campaign_name") for b in brands.values() if b.get("campaign_name")]
    brand_keys = list(brands)
    vds = list(_yaml("video_dominations.yaml").get("options", {}))
    takeovers = list(_yaml("operative_takeovers.yaml").get("types", {}))
    sources = {
        "region": regions,
        "language": ["English", "French"],
        "campaign name": campaigns,
        "content type": ["show", "movie"],
        "video domination": vds,
        "takeover": takeovers,
        "brand": brand_keys,
    }
    # Every Products toggle shares one Yes/No list.
    for label in _product_toggle_labels():
        sources[label] = ["Yes", "No"]
    return sources


# --- palette -------------------------------------------------------------- #
NAVY, BLUE = "0B3D91", "1F6FEB"
YELLOW = "FFF7D6"        # required-input value cells
GREY_FILL = "EDEFF4"     # auto-derived / override rows (leave blank)
BAND = "F5F8FD"          # subtle row band
BORDER_C = "D5DCE8"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=BLUE)
TITLE_FILL = PatternFill("solid", fgColor=NAVY)
REQ_FILL = PatternFill("solid", fgColor=YELLOW)
AUTO_FILL = PatternFill("solid", fgColor=GREY_FILL)
BAND_FILL = PatternFill("solid", fgColor=BAND)
WHITE_BOLD = Font(color="FFFFFF", bold=True)
WHITE_BOLD_LG = Font(color="FFFFFF", bold=True, size=15)
GREY = Font(color="7A7A7A")
GREY_IT = Font(color="8A8A8A", italic=True)
FIELD_FONT = Font(bold=True, color="21324D")
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_MID = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")
_thin = Side(style="thin", color=BORDER_C)
BOX = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _title_banner(ws, span_cols: int, title: str, subtitle: str) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span_cols)
    t = ws.cell(row=1, column=1, value=title)
    t.fill = TITLE_FILL; t.font = WHITE_BOLD_LG; t.alignment = CENTER
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span_cols)
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(color="44506A", italic=True); s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18


def build(out: Path | None = None) -> Path:
    wb = Workbook()
    info_ws = wb.active
    info_ws.title = "Instructions"
    plan_ws = wb.create_sheet("Plan")
    targ_ws = wb.create_sheet("Targeting")
    lists_ws = wb.create_sheet("_Lists")

    sources = dict(dropdown_sources())
    sources["kids audience"] = ["older", "younger"]   # Targeting-tab dropdown

    # --- hidden _Lists sheet: one column per distinct value-set (Yes/No is shared) ---
    col_ranges: dict[str, str] = {}
    range_by_values: dict[tuple, str] = {}
    next_col = 1
    for label, values in sources.items():
        key = tuple(values)
        if key not in range_by_values:
            letter = get_column_letter(next_col)
            lists_ws.cell(row=1, column=next_col, value=label)
            for r, v in enumerate(values, start=2):
                lists_ws.cell(row=r, column=next_col, value=v)
            range_by_values[key] = f"_Lists!${letter}$2:${letter}${1 + len(values)}"
            next_col += 1
        col_ranges[label] = range_by_values[key]
    lists_ws.sheet_state = "hidden"

    def _dropdown(ws, cell, rng, note):
        dv = DataValidation(type="list", formula1=rng, allow_blank=True,
                            showErrorMessage=True, showInputMessage=True)
        dv.error = "Please pick a value from the dropdown list."
        dv.errorTitle = "Choose from the list"
        dv.promptTitle = "Pick from the list ▾"
        dv.prompt = (note or "Choose a value.")[:250]
        ws.add_data_validation(dv)
        dv.add(cell)

    # --- Instructions tab ---
    info_ws.sheet_view.showGridLines = False
    info_ws.column_dimensions["A"].width = 3
    info_ws.column_dimensions["B"].width = 100
    _title_banner(info_ws, 3, "Paramount Promo — Campaign Plan",
                  "One workbook per campaign. Fill the Plan tab, then the Targeting tab.")
    guide = [
        ("h", "How to use"),
        ("b", "1.  Make a copy of this workbook for each campaign."),
        ("b", "2.  On the Plan tab, fill the highlighted (yellow) cells. Use the dropdowns where you see the ▾ arrow — they prevent typos."),
        ("b", "3.  On the Targeting tab, list your Networks / Genres / Showlist / Pluto Categories / Pluto Channels — one item per row, down each column."),
        ("b", "4.  Leave the grey “auto-derived” fields blank — the automation fills them from the Campaign you pick."),
        ("b", "5.  Attach the workbook (or the Targeting tab) to the Salesforce Case and set Status = “Ready for Automation.”"),
        ("sp", ""),
        ("h", "Colour key"),
        ("req", "Yellow  =  fill this in (required / your input)"),
        ("auto", "Grey  =  leave blank (auto-derived from the Campaign)"),
        ("sec", "Blue band  =  a section header"),
        ("sp", ""),
        ("h", "Tips"),
        ("b", "•  Pick the Campaign first — it sets the Brand, Advertiser, and default products for you."),
        ("b", "•  Pluto category / channel names differ by region. Use your region's real names (ask Ad Ops for the per-region list)."),
        ("b", "•  Products are Yes/No toggles — leave blank to keep the brand's standard set; set No to drop one; Yes to add one."),
    ]
    r = 4
    for kind, text in guide:
        c = info_ws.cell(row=r, column=2, value=text)
        if kind == "h":
            c.font = Font(bold=True, size=12, color=NAVY)
        elif kind == "req":
            c.fill = REQ_FILL; c.font = Font(bold=True); c.border = BOX
        elif kind == "auto":
            c.fill = AUTO_FILL; c.font = GREY_IT; c.border = BOX
        elif kind == "sec":
            c.fill = SECTION_FILL; c.font = WHITE_BOLD
        else:
            c.font = Font(color="333333")
        c.alignment = WRAP_MID
        info_ws.row_dimensions[r].height = 20 if kind != "sp" else 8
        r += 1

    # --- Plan tab ---
    plan_ws.sheet_view.showGridLines = False
    rows = _csv_rows("Plan.csv")
    plan_ws.column_dimensions["A"].width = 28
    plan_ws.column_dimensions["B"].width = 34
    plan_ws.column_dimensions["C"].width = 66
    _title_banner(plan_ws, 3, "Plan — campaign details",
                  "Fill the yellow cells. Dropdowns (▾) prevent typos. Grey = leave blank (auto-filled).")
    HEADROW = 3
    for c, head in enumerate(["Field", "Your entry", "Notes / allowed values"], start=1):
        cell = plan_ws.cell(row=HEADROW, column=c, value=head)
        cell.fill = HEADER_FILL; cell.font = WHITE_BOLD; cell.border = BOX
        cell.alignment = CENTER if c < 3 else Alignment(vertical="center")
    plan_ws.row_dimensions[HEADROW].height = 20

    in_override = False
    band = False
    for i, row in enumerate(rows[1:]):
        rr = HEADROW + 1 + i
        field = row[0] if len(row) > 0 else ""
        value = row[1] if len(row) > 1 else ""
        note = row[2] if len(row) > 2 else ""
        is_section = field.startswith("—")
        fcell = plan_ws.cell(row=rr, column=1, value=field)
        vcell = plan_ws.cell(row=rr, column=2, value=value)
        ncell = plan_ws.cell(row=rr, column=3, value=note)
        for cc in (fcell, vcell, ncell):
            cc.border = BOX
        ncell.font = GREY; ncell.alignment = WRAP
        plan_ws.row_dimensions[rr].height = 17

        if is_section:
            in_override = "OVERRIDE" in field.upper()
            plan_ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=3)
            for c in (1, 2, 3):
                sc = plan_ws.cell(row=rr, column=c)
                sc.fill = SECTION_FILL; sc.font = WHITE_BOLD; sc.border = BOX
            plan_ws.cell(row=rr, column=1).value = field.strip("— ").strip()
            plan_ws.cell(row=rr, column=1).alignment = Alignment(vertical="center")
            plan_ws.row_dimensions[rr].height = 19
            band = False
            continue

        fcell.font = FIELD_FONT
        band = not band
        if band:
            fcell.fill = BAND_FILL; ncell.fill = BAND_FILL
        key = field.strip().lower()
        has_dropdown = key in col_ranges
        required = "(required)" in note.lower()
        if in_override:
            fcell.fill = AUTO_FILL; vcell.fill = AUTO_FILL
            fcell.font = GREY_IT
        elif required or has_dropdown:
            vcell.fill = REQ_FILL     # invite input
        if has_dropdown:
            _dropdown(plan_ws, vcell, col_ranges[key], note)
    plan_ws.freeze_panes = f"A{HEADROW + 1}"

    # --- Targeting tab ---
    targ_ws.sheet_view.showGridLines = False
    trows = _csv_rows("Targeting.csv")
    _title_banner(targ_ws, len(trows[0]), "Targeting — one item per row, down each column",
                  "List everything to target. Audience Segments (Tier 1) auto-resolve from the Showlist — leave blank.")
    HR = 3
    kids_col = None
    for c, head in enumerate(trows[0], start=1):
        cell = targ_ws.cell(row=HR, column=c, value=head)
        cell.fill = HEADER_FILL; cell.font = WHITE_BOLD; cell.alignment = WRAP; cell.border = BOX
        targ_ws.column_dimensions[get_column_letter(c)].width = 26
        if head.strip().lower().startswith("kids audience"):
            kids_col = c
    targ_ws.row_dimensions[HR].height = 34
    for i, row in enumerate(trows[1:]):
        rr = HR + 1 + i
        for c in range(1, len(trows[0]) + 1):
            val = row[c - 1] if c - 1 < len(row) else ""
            cell = targ_ws.cell(row=rr, column=c, value=val)
            cell.border = BOX; cell.alignment = WRAP
            if rr % 2 == 0:
                cell.fill = BAND_FILL
    # Kids Audience dropdown down its column (Kids brands only).
    if kids_col:
        rng = col_ranges["kids audience"]
        for rr in range(HR + 1, HR + 60):
            _dropdown(targ_ws, targ_ws.cell(row=rr, column=kids_col), rng,
                      "older / younger (Kids brands only)")
    targ_ws.freeze_panes = f"A{HR + 1}"

    out = out or (TEMPLATE_DIR / "Campaign-Plan-Template.xlsx")
    wb.save(out)
    return out


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
